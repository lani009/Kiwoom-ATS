import openpyxl


class ConfigParser:
    FILE_PATH: str
    __row_start: int
    __row_end: int

    def __init__(self):
        self.FILE_PATH = "./resources/config/config_stock.xlsx"
        self.__row_start = 9
        self.__row_end = 28

    @classmethod
    def __get_instance(cls):
        return cls.__instance

    @classmethod
    def instance(cls, *args, **kargs):
        cls.__instance = cls(*args, **kargs)
        cls.instance = cls.__get_instance
        return cls.__instance

    def load_stock_config(self):
        """로컬에서 주식 설정을 불러온다."""
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["main"]
        config = list()
        for i in range(self.__row_start, self.__row_end):
            stock_name = sheet.cell(i, 2).value
            stock_code = sheet.cell(i, 3).value
            if stock_code is None:
                continue

            data = {
                "stock_code": stock_code,
                "stock_name": stock_name,
                "B1": {"price": sheet.cell(i, 4).value, "qty": sheet.cell(i, 5).value},
                "B2": {"price": sheet.cell(i, 6).value, "qty": sheet.cell(i, 7).value},
                "S1": {"price": sheet.cell(i, 8).value, "qty": sheet.cell(i, 9).value},
                "S2": {"price": sheet.cell(i, 10).value, "qty": sheet.cell(i, 11).value},
                "S3": {"price": sheet.cell(i, 12).value, "qty": sheet.cell(i, 13).value},
                "S4": {"price": sheet.cell(i, 14).value, "qty": sheet.cell(i, 15).value},
                "S5": {"price": sheet.cell(i, 16).value, "qty": sheet.cell(i, 17).value},
            }

            config.append(data)
        wb.close()
        return config

    def add_unfinished_stock(self, data_list):
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["trading"]

        for i in range(self.__row_start, 30):
            for j in range(2, 18):
                sheet.cell(i, j).value = None

        i = self.__row_start
        for data in data_list:
            sheet.cell(i, 2).value = data["stock_name"]
            sheet.cell(i, 3).value = data["stock_code"]
            sheet.cell(i, 4).value = data["B1"]["price"]
            sheet.cell(i, 5).value = data["B1"]["qty"]
            sheet.cell(i, 6).value = data["B2"]["price"]
            sheet.cell(i, 7).value = data["B2"]["qty"]
            sheet.cell(i, 8).value = data["S1"]["price"]
            sheet.cell(i, 9).value = data["S1"]["qty"]
            sheet.cell(i, 10).value = data["S2"]["price"]
            sheet.cell(i, 11).value = data["S2"]["qty"]
            sheet.cell(i, 12).value = data["S3"]["price"]
            sheet.cell(i, 13).value = data["S3"]["qty"]
            sheet.cell(i, 14).value = data["S4"]["price"]
            sheet.cell(i, 15).value = data["S4"]["qty"]
            sheet.cell(i, 16).value = data["S5"]["price"]
            sheet.cell(i, 17).value = data["S5"]["qty"]
            sheet.cell(i, 18).value = data["state"]
            i += 1
        wb.save(self.FILE_PATH)
        wb.close()

    def remove_stock_config(self, stock_code: str):
        row_index = self.find_stock_row(stock_code, "main")
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["main"]

        for j in range(2, 19):
            sheet.cell(row_index, j).value = None

        wb.save(self.FILE_PATH)
        wb.close()

    def find_stock_row(self, stock_code: str, sheet: str) -> int:
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb[sheet]

        for i in range(self.__row_start, self.__row_end):
            if sheet.cell(i, 3).value == stock_code:
                wb.close()
                return i
        wb.close()
        raise KeyError(f"No such key {stock_code} in config_stock.xlsx")

    def load_maximum_trading(self):
        """최대 거래 가능 종목 수

        Returns
        -------
        int:
            종목 수
        """
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["setting"]

        val = int(sheet["D5"].value)

        wb.close()
        return val

    def load_is_power_off(self):
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["setting"]

        val = sheet["H5"].value.strip() == "y"

        wb.close()
        return val

    def get_account_number(self):
        wb = openpyxl.load_workbook(self.FILE_PATH)
        sheet = wb["setting"]

        val = sheet["D9"].value.strip()

        wb.close()
        return val
